"""Excel String Extractor"""

# exsextractor.py


import json
import os
import sys
from enum import Enum, auto
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import argparse
from .args import parse_args
from .config import ConfigType
from .config import config_align
from .config import config_2align
from .config import check_config_list_config


CONF_TEMPLATE_JSON_PATH = os.path.join('exsextractor', 'template', 'conf.template.json')


def get_input_from_list_file (data: list|dict, config: dict):
    if isinstance(data, dict):
        list_sheet = None
        if 'list' in data and 'sheets' in data['list'] and 'list_sheet' in data['list']['sheets']:
            list_sheet = data['list']['sheets']['list_sheet']
        if ('list' in data and 'file_name' in data['list']):
            if data['list']['file_name'].endswith('.xlsx'):
                return get_input_from_file_list_xlsx(data['list']['file_name'], list_sheet)
            if data['list']['file_name'].endswith('.csv'):
                return get_input_from_file_list_csv(data['list']['file_name'], config['delimiter'])

    elif data[0].endswith('.xlsx'):
        return get_input_from_file_list_xlsx(data[0], data[1] if len(data) > 1 else None)
    elif data[0].endswith('.csv'):
        return get_input_from_file_list_csv(data[0], config['delimiter'])
    print("File type not supported")
    print(data)
    

def get_input_from_file_list_xlsx (file_name: str, list_sheet: str|None = None):
    input_files = []
    # TODO    
    return input_files


def get_input_from_file_list_csv (file_name: str, delimiter: str = ';'):
    input_files = []
    # TODO    
    return input_files


def get_config_from_sheet (list_data):
    config_2 = {}
    
    # TODO
    
    return config_2


def get_config_from_file (file_name, config_sheet = None):
    if (file_name.endswith('.json')):
        return get_config_from_file_json(file_name)
    elif (file_name.endswith('.xlsx')):
        return get_config_from_file_xlsx(file_name, config_sheet)
    raise Exception('file type not supported')


def get_config_from_file_json (file_name):
    with open(file_name, 'r') as file:
        return json.load(file)


# TODO deve controllare anche tutte le directory passate in config?
def get_config_from_file_xlsx (file_name, config_sheet = None):
    wb = load_workbook(file_name, data_only=True)
    if config_sheet is None:
        ws = wb.worksheets[0]
    else:
        ws = wb[config_sheet]

    # Trova la colonna valorizzata
    value_column_index = 0 # values_only = True quindi row è una lista 0-based
    for cell in ws[1]:
        if cell.value == 'x':
            break
        value_column_index += 1

    config = {}

    row_num = 2
    for row in ws.iter_rows(min_row = 2, max_row = ws.max_row, min_col = 1, values_only = True):
        if row[0] == 'input':
            if row[value_column_index] != None:
                if not 'input' in config:
                    config['input'] = []
                config['input'].extend(row[value_column_index])
        elif row[0] == 'label_name' and row[1] == 'name':
            next_row = [cell.value for cell in ws[row_num + 1]]
            if next_row[0] != 'label_name' or next_row[1] != 'label':
                raise ValueError("Errore: la riga successiva a 'label_name' deve contenere 'label_name' e 'label'")
            if not 'label_name' in config:
                config['label_name'] = []
            column = value_column_index
            while row[column] != None:
                name = row[column]
                label = next_row[column]
                config['label_name'].append((name,label))
                column += 1
        row_num += 1
    print("\nget_config_from_file_xlsx:")
    print(config)
    return config


def main ():
    (parser, args) = parse_args()
    # print(args)


    # configurazione
    with open(CONF_TEMPLATE_JSON_PATH, 'r') as file:
        config = json.load(file)

    config_align(config, args)

    if (args.check('--config')):
        config_2 = get_config_from_file(args.config[0])
        if (args.check('--input') or args.check('--list') or args.check('--list-config')):
            if (args.check('list-config')):
                parser.print_exit('''argument -co/--config: not allowed with argument -lc/--list-config''')
            try:
                check_config_list_config(config_2)
            except Exception as e:
                parser.print_exit(e)
            config_2align(config, config_2, ConfigType.INPUT if args.check('input') else ConfigType.LIST)
        else:
            # è stato passato soltanto un file di configurazione
            if ('input' in config_2 and 'list' in config_2):
                parser.print_exit('''config must contain either "input", "list" or "list_config"''')
            config_2align(config, config_2, ConfigType.CONFIG)

    if (args.check('--list-config')):
        # assumo che non sia stato passato un file di configurazione separato e che
        # non siano stati passati --input e --list
        config_2 = get_config_from_sheet(args.list_config)
        try:
            check_config_list_config(config_2)
        except Exception as e:
            parser.print_exit(e)
        config_2align(config, config_2, ConfigType.LIST_CONFIG)
    
    
    # lista dei file di input
    input_files = []
    if (args.check('--input')):
        input_files.extend(args.input)
    if (args.check('--list')):
        input_files.extend(get_input_from_list_file(args.list, config))
    if (args.check('--list-config')):
        input_files.extend(get_input_from_list_file(args.list_config, config))
    if ('input' in config):
        input_files.extend(config['input'])
    if ('list' in config and 'file_name' in config['list'] and config['list']['file_name'] != None):
        input_files.extend(get_input_from_list_file(config['list'], config))
    if ('list_config' in config and 'file_name' in config['list_config'] and config['list_config']['file_name'] != None):
        input_files.extend(get_input_from_list_file(config['list_config'], config))

    print("\n", config)
    print("\n", input_files)