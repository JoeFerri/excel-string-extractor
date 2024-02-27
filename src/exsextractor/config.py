"""Config for exsextractor"""

# config.py


import json
from enum import Enum, auto
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import argparse

from .args import parse_args


class ConfigType(Enum):
    INPUT = auto()
    LIST = auto()
    LIST_CONFIG = auto()
    CONFIG = auto()


def get_config_by_template (path):
    with open(path, 'r') as file:
        return json.load(file)


def check_config_list_config (config_2):
    # TODO cosa volevo controllare?
    return True
    # if (isinstance(config_2, object)
    #     and ('input' in config_2
    #     or 'list' in config_2)):
    #     raise Exception('config_2 must contain either "input", "list" or "list_config"')



def get_config_from_sheet (list_data):
    config_2 = {}
    
    # TODO
    
    return config_2


def get_config_from_file (file_name, config_sheet = None):
    if (file_name.endswith('.json')):
        return get_config_from_file_json(file_name)
    elif (file_name.endswith('.xlsx')):
        return get_config_from_file_xlsx(file_name, config_sheet)
    raise Exception('file type not supported')


def get_config_from_file_json (file_name):
    with open(file_name, 'r') as file:
        return json.load(file)


# TODO deve controllare anche tutte le directory passate in config?
def get_config_from_file_xlsx (file_name, config_sheet = None):
    wb = load_workbook(file_name, data_only=True)
    if config_sheet is None:
        ws = wb.worksheets[0]
    else:
        ws = wb[config_sheet]

    # Trova la colonna valorizzata
    value_column_index = 0 # values_only = True ~> row[index] (0-based)
    # ws[index] (1-based)
    for cell in ws[1]:
        if cell.value == 'x':
            break
        value_column_index += 1

    config = {}

    row_num = 2 # inizio la lettura dalla riga 2 - ws[row_num] (1-based)
    for row in ws.iter_rows(min_row = 2, max_row = ws.max_row, min_col = 1, values_only = True):
        # row[index] (0-based)
        if row[0] == 'input':
            if row[value_column_index] != None:
                config['input'] = row[value_column_index]

        elif row[0] == 'label_name' and row[1] == 'name':
            # Trova la riga successiva di soli valori
            # next_row[index] (0-based)
            next_row = [cell.value for cell in ws[row_num + 1]]
            if next_row[0] != 'label_name' or next_row[1] != 'label':
                raise ValueError("Errore: la riga successiva a 'label_name' deve contenere 'label_name' e 'label'")
            if not 'label_name' in config:
                config['label_name'] = []
            column = value_column_index
            max_column_index = max(len(row), len(next_row)) -1 # (0-based)
            # scansiona le celle a destra della colonna valorizzata
            while column <= max_column_index:
                name = row[column]
                label = next_row[column]
                if not name is None and not label is None:
                    config['label_name'].append((name,label))
                elif (name is None and not label is None) or (not name is None and label is None):
                    raise ValueError("Errore: 'label_name' deve avere nome e label")
                # nel caso name is None and label is None non fa niente
                column += 1

        if row[0] == 'unique':
            if row[value_column_index] != None:
                config['unique'] = row[value_column_index]
        # TODO

        row_num += 1
    print("\nget_config_from_file_xlsx:")
    print(config)
    return config


def config_align (config: dict, args: argparse.Namespace):
    config['label_name'].extend(args.label_name)
    
    config['unique'] = args.unique
    config['unique_file'].extend(args.unique_file)
    config['unique_sheet'].extend(args.unique_sheet)
    config['unique_count'] = args.unique_count
    
    config['out'] = args.out[0] if len(args.out) > 0 else config['out']
    config['format_out'] = args.format_out
    config['delimiter'] = args.delimiter
    config['multi_file'] = args.multi_file
    config['multi_proc'] = args.multi_proc
    config['multi_thr'] = args.multi_thr
    config['one_column_capture'] = args.one_column_capture
    
    config['value'] = args.value
    config['digit'] = args.digit
    config['word'] = args.word
    config['text'] = args.text
    config['string'] = args.string
    config['raw'] = args.raw
    config['regex'] = args.regex
    
    config['include_column'].extend(args.include_column)
    config['exclude_column'].extend(args.exclude_column)
    config['rename'].extend(args.rename)
    config['custom'].extend(args.custom)
    config['sequence'] = args.sequence
    
    config['input_directory'].extend(args.input_directory)
    config['output_directory'] = args.output_directory
    return config


def config_2align (config: dict, config_2: dict, config_type: ConfigType):
    if 'label_name' in config_2: config['label_name'].extend(config_2['label_name'])
    
    if 'unique' in config_2: config['unique'] = config_2['unique']
    if 'unique_file' in config_2: config['unique_file'].extend(config_2['unique_file'])
    if 'unique_sheet' in config_2: config['unique_sheet'].extend(config_2['unique_sheet'])
    if 'unique_count' in config_2: config['unique_count'] = config_2['unique_count']
    
    if config_type == ConfigType.CONFIG:
        if 'input' in config_2: config['input'] = config_2['input']
        if 'list' in config_2 and 'file_name' in config_2['list']: config['list']['file_name'] = config_2['list']['file_name']
        if 'list' in config_2 and 'sheets' in config_2['list'] and 'list_sheet' in config_2['list']['sheets']: config['list']['sheets']['list_sheet'] = config_2['list']['sheets']['list_sheet']
        if 'list' in config_2 and 'sheets' in config_2['list'] and 'config_sheet' in config_2['list']['sheets']: config['list']['sheets']['config_sheet'] = config_2['list']['sheets']['config_sheet']
    
    if 'out' in config_2: config['out'] = config_2['out']
    if 'format_out' in config_2: config['format_out'] = config_2['format_out']
    if 'delimiter' in config_2: config['delimiter'] = config_2['delimiter']
    if 'multi_file' in config_2: config['multi_file'] = config_2['multi_file']
    if 'multi_proc' in config_2: config['multi_proc'] = config_2['multi_proc']
    if 'multi_thr' in config_2: config['multi_thr'] = config_2['multi_thr']
    if 'one_column_capture' in config_2: config['one_column_capture'] = config_2['one_column_capture']
    
    if 'value' in config_2: config['value'] = config_2['value']
    if 'digit' in config_2: config['digit'] = config_2['digit']
    if 'word' in config_2: config['word'] = config_2['word']
    if 'text' in config_2: config['text'] = config_2['text']
    if 'string' in config_2: config['string'] = config_2['string']
    if 'raw' in config_2: config['raw'] = config_2['raw']
    if 'regex' in config_2: config['regex'] = config_2['regex']
    
    if 'include_column' in config_2: config['include_column'].extend(config_2['include_column'])
    if 'exclude_column' in config_2: config['exclude_column'].extend(config_2['exclude_column'])
    if 'rename' in config_2: config['rename'].extend(config_2['rename'])
    if 'custom' in config_2: config['custom'].extend(config_2['custom'])
    if 'sequence' in config_2: config['sequence'] = config_2['sequence']
    
    if 'input_directory' in config_2: config['input_directory'].extend(config_2['input_directory'])
    if 'output_directory' in config_2: config['output_directory'] = config_2['output_directory']
    return config